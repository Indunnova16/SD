"""Tests for import/export masivo con `rol_acceso` + `supervisor_documento`
(SD#58, sub-item A3).

Cubre:
  - `BulkUploadService.ROL_MAP` / `ROL_SUGERIDO_POR_JOB_PROFILE_CODE`:
    fila con `rol_acceso` explícito (gana sobre la sugerencia), fila sin la
    columna cae al mismo mapeo sugerido de A1/A2 (7 códigos confiables),
    fila con `job_profile` en el bucket "sin sugerencia" (CAPATAZ) y sin
    `rol_acceso` -> fila rechazada con error listado (misma regla de A2:
    "aplicar sugerencia automática solo si el job_profile tiene mapeo
    confiable, exigir rol explícito si no").
  - `rol_acceso` con valor no reconocido (typo) -> error listado, no se
    adivina.
  - `supervisor_documento`: asignación válida contra un usuario YA EXISTENTE
    (dato legacy, no creado en el mismo batch) con rol Coordinador/Admin;
    documento inexistente -> error; supervisor con rol Ejecutor -> error
    (mismo queryset que restringe el selector de A2); auto-supervisión
    (mismo documento en la fila) -> error (misma regla de A1 `clean()`).
  - `BulkUploadService.generate_template()` incluye las 2 columnas nuevas.
  - `ExportService.export_pending_users()` incluye las 2 columnas nuevas con
    los valores reales de rol/supervisor del usuario.
"""

from datetime import date

from django.contrib.auth import get_user_model
from django.test import TestCase

from apps.accounts.services import BulkUploadService, ExportService
from apps.courses.models import Category, Course, Enrollment, JobProfileType

User = get_user_model()


def _row(row_num=2, **overrides):
    """Fila mínima válida de bulk upload (dict como lo produce `parse_excel`).

    Los tests sobreescriben solo los campos relevantes al caso bajo prueba.
    """
    data = {
        "_row_num": row_num,
        "nombre": "Nuevo",
        "apellido": "Usuario",
        "numero_documento": "600000001",
        "tipo_documento": "CC",
        "correo": "",
        "telefono": "",
        "cargo": "Cargo de prueba",
        "perfil_ocupacional": "LINIERO",
        "tipo_vinculacion": "directo",
        "fecha_ingreso": date(2026, 1, 1),
        "estado": "activo",
    }
    data.update(overrides)
    return data


class BulkImportRolExplicitoTests(TestCase):
    """`rol_acceso` explícito en la fila gana sobre la sugerencia del
    `job_profile` (misma jerarquía que A2: valor explícito > sugerencia)."""

    def setUp(self):
        self.jp_liniero, _ = JobProfileType.objects.get_or_create(
            code="LINIERO", defaults={"name": "Liniero", "order": 1}
        )

    def test_rol_acceso_explicito_sobreescribe_sugerencia(self):
        # LINIERO sugeriría EJECUTOR, pero la fila pide COORDINADOR explícito.
        row = _row(rol_acceso="COORDINADOR")
        created, errors = BulkUploadService.create_users_from_rows([row])

        self.assertEqual(errors, [])
        self.assertEqual(len(created), 1)
        self.assertEqual(created[0].rol, "COORDINADOR")

    def test_rol_acceso_no_distingue_mayusculas(self):
        row = _row(rol_acceso="administrador")
        created, errors = BulkUploadService.create_users_from_rows([row])

        self.assertEqual(errors, [])
        self.assertEqual(created[0].rol, "ADMINISTRADOR")

    def test_rol_acceso_valor_no_reconocido_es_error(self):
        row = _row(rol_acceso="MANAGER")
        created, errors = BulkUploadService.create_users_from_rows([row])

        self.assertEqual(created, [])
        self.assertEqual(len(errors), 1)
        self.assertIn("rol_acceso", errors[0])
        self.assertFalse(User.objects.filter(document_number=row["numero_documento"]).exists())


class BulkImportRolSugeridoTests(TestCase):
    """Fila SIN columna `rol_acceso` cae al mismo mapeo sugerido de A1/A2
    para los 7 códigos confiables."""

    def setUp(self):
        self.jp_liniero, _ = JobProfileType.objects.get_or_create(
            code="LINIERO", defaults={"name": "Liniero", "order": 1}
        )
        self.jp_coord_hseq, _ = JobProfileType.objects.get_or_create(
            code="COORDINADOR_HSEQ", defaults={"name": "Coordinador HSEQ", "order": 2}
        )
        self.jp_admin, _ = JobProfileType.objects.get_or_create(
            code="ADMINISTRADOR", defaults={"name": "Administrador", "order": 3}
        )

    def test_liniero_sugiere_ejecutor(self):
        row = _row(numero_documento="600000010", perfil_ocupacional="LINIERO")
        created, errors = BulkUploadService.create_users_from_rows([row])

        self.assertEqual(errors, [])
        self.assertEqual(created[0].rol, "EJECUTOR")

    def test_coordinador_hseq_sugiere_coordinador(self):
        row = _row(
            numero_documento="600000011",
            perfil_ocupacional="coordinador hseq",
        )
        created, errors = BulkUploadService.create_users_from_rows([row])

        self.assertEqual(errors, [])
        self.assertEqual(created[0].rol, "COORDINADOR")

    def test_administrador_sugiere_administrador(self):
        row = _row(numero_documento="600000012", perfil_ocupacional="administrador")
        created, errors = BulkUploadService.create_users_from_rows([row])

        self.assertEqual(errors, [])
        self.assertEqual(created[0].rol, "ADMINISTRADOR")


class BulkImportRolSinSugerenciaTests(TestCase):
    """`job_profile` en el bucket "sin sugerencia" (evidencia F2: CAPATAZ,
    "TODOS LOS CARGOS", CONTRATISTA, CONDUCTOR, COORDINADOR_VIZ) exige
    `rol_acceso` explícito — sin él, la fila se rechaza (no se adivina)."""

    def setUp(self):
        self.jp_capataz = JobProfileType.objects.create(code="CAPATAZ", name="Capataz", order=20)

    def test_capataz_sin_rol_acceso_es_error_listado(self):
        row = _row(perfil_ocupacional="CAPATAZ")
        created, errors = BulkUploadService.create_users_from_rows([row])

        self.assertEqual(created, [])
        self.assertEqual(len(errors), 1)
        self.assertIn("CAPATAZ", errors[0])
        self.assertIn("rol_acceso", errors[0])
        self.assertFalse(User.objects.filter(document_number=row["numero_documento"]).exists())

    def test_capataz_con_rol_acceso_explicito_es_valido(self):
        row = _row(perfil_ocupacional="CAPATAZ", rol_acceso="COORDINADOR")
        created, errors = BulkUploadService.create_users_from_rows([row])

        self.assertEqual(errors, [])
        self.assertEqual(len(created), 1)
        self.assertEqual(created[0].rol, "COORDINADOR")
        self.assertEqual(created[0].job_profile.code, "CAPATAZ")

    def test_perfil_ocupacional_pasa_codigo_literal_sin_alias(self):
        """Sin alias "amigable" en JOB_PROFILE_MAP para CAPATAZ, el código
        literal ('CAPATAZ', tal como está en `job_profile_types.code`)
        igual debe resolver al JobProfileType real — no degradar a LINIERO
        en silencio (issue #58 A3: antes de este fix, cualquier perfil sin
        alias caía siempre a LINIERO)."""
        row = _row(perfil_ocupacional="CAPATAZ", rol_acceso="EJECUTOR")
        created, errors = BulkUploadService.create_users_from_rows([row])

        self.assertEqual(errors, [])
        self.assertEqual(created[0].job_profile_id, self.jp_capataz.pk)


class BulkImportSupervisorTests(TestCase):
    """`supervisor_documento`: asignación contra un usuario ya existente en
    BD (dato legacy — no creado en el mismo lote de importación)."""

    def setUp(self):
        self.jp_liniero, _ = JobProfileType.objects.get_or_create(
            code="LINIERO", defaults={"name": "Liniero", "order": 1}
        )
        self.jp_coord_hseq, _ = JobProfileType.objects.get_or_create(
            code="COORDINADOR_HSEQ", defaults={"name": "Coordinador HSEQ", "order": 2}
        )
        # Usuario "legacy": ya existía en BD antes de correr este import,
        # con rol Coordinador (candidato válido a supervisor).
        self.coordinador_legacy = User.objects.create_user(
            email="coord_legacy_a3@example.com",
            password="x",
            first_name="Coordinador",
            last_name="Legacy",
            document_type="CC",
            document_number="600000100",
            job_position="Coordinador HSEQ",
            job_profile=self.jp_coord_hseq,
            hire_date=date(2023, 1, 1),
            rol="COORDINADOR",
        )
        self.ejecutor_legacy = User.objects.create_user(
            email="ejec_legacy_a3@example.com",
            password="x",
            first_name="Ejecutor",
            last_name="Legacy",
            document_type="CC",
            document_number="600000101",
            job_position="Liniero",
            job_profile=self.jp_liniero,
            hire_date=date(2023, 1, 1),
            rol="EJECUTOR",
        )

    def test_supervisor_documento_valido_se_asigna(self):
        row = _row(
            numero_documento="600000110",
            supervisor_documento="600000100",  # coordinador_legacy
        )
        created, errors = BulkUploadService.create_users_from_rows([row])

        self.assertEqual(errors, [])
        self.assertEqual(len(created), 1)
        self.assertEqual(created[0].supervisor_id, self.coordinador_legacy.pk)

    def test_supervisor_documento_vacio_queda_sin_supervisor(self):
        row = _row(numero_documento="600000111", supervisor_documento="")
        created, errors = BulkUploadService.create_users_from_rows([row])

        self.assertEqual(errors, [])
        self.assertIsNone(created[0].supervisor_id)

    def test_supervisor_documento_inexistente_es_error(self):
        row = _row(numero_documento="600000112", supervisor_documento="999999999")
        created, errors = BulkUploadService.create_users_from_rows([row])

        self.assertEqual(created, [])
        self.assertEqual(len(errors), 1)
        self.assertIn("999999999", errors[0])
        self.assertFalse(User.objects.filter(document_number="600000112").exists())

    def test_supervisor_con_rol_ejecutor_es_error(self):
        """Misma regla del queryset de A2 (`supervisor` limitado a
        rol in [COORDINADOR, ADMINISTRADOR]) — un Ejecutor no puede ser
        supervisor, ni siquiera vía import masivo."""
        row = _row(
            numero_documento="600000113",
            supervisor_documento="600000101",  # ejecutor_legacy
        )
        created, errors = BulkUploadService.create_users_from_rows([row])

        self.assertEqual(created, [])
        self.assertEqual(len(errors), 1)
        self.assertIn("600000101", errors[0])
        self.assertFalse(User.objects.filter(document_number="600000113").exists())

    def test_auto_supervision_es_error(self):
        """`supervisor_documento` igual a `numero_documento` de la misma
        fila (misma regla de `User.clean()` en A1: rechazar auto-referencia)."""
        row = _row(
            numero_documento="600000114",
            supervisor_documento="600000114",
        )
        created, errors = BulkUploadService.create_users_from_rows([row])

        self.assertEqual(created, [])
        self.assertEqual(len(errors), 1)
        self.assertIn("sí mismo", errors[0])
        self.assertFalse(User.objects.filter(document_number="600000114").exists())


class GenerateTemplateColumnsTests(TestCase):
    """`generate_template()` incluye las 2 columnas nuevas en el header."""

    def test_template_incluye_rol_acceso_y_supervisor_documento(self):
        content = BulkUploadService.generate_template()
        self.assertTrue(content)

        from io import BytesIO

        import openpyxl

        wb = openpyxl.load_workbook(BytesIO(content))
        ws = wb["Usuarios"]
        headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]

        self.assertIn("rol_acceso", headers)
        self.assertIn("supervisor_documento", headers)


class ExportPendingUsersColumnsTests(TestCase):
    """`ExportService.export_pending_users()` incluye las 2 columnas nuevas
    con los valores reales del usuario, probado contra un usuario/curso
    "legacy" ya existente en BD (no una fixture nueva de este test)."""

    def setUp(self):
        self.jp_liniero, _ = JobProfileType.objects.get_or_create(
            code="LINIERO", defaults={"name": "Liniero", "order": 1}
        )
        self.jp_coord_hseq, _ = JobProfileType.objects.get_or_create(
            code="COORDINADOR_HSEQ", defaults={"name": "Coordinador HSEQ", "order": 2}
        )
        self.coordinador = User.objects.create_user(
            email="coord_export_a3@example.com",
            password="x",
            first_name="Coordinador",
            last_name="Export",
            document_type="CC",
            document_number="600000200",
            job_position="Coordinador HSEQ",
            job_profile=self.jp_coord_hseq,
            hire_date=date(2023, 1, 1),
            rol="COORDINADOR",
        )
        self.ejecutor = User.objects.create_user(
            email="ejec_export_a3@example.com",
            password="x",
            first_name="Ejecutor",
            last_name="Export",
            document_type="CC",
            document_number="600000201",
            job_position="Liniero",
            job_profile=self.jp_liniero,
            hire_date=date(2023, 1, 1),
            rol="EJECUTOR",
            supervisor=self.coordinador,
        )
        self.category = Category.objects.create(name="Seguridad", slug="seguridad-a3")
        self.course = Course.objects.create(
            code="CUR-A3-001",
            title="Curso Pendiente A3",
            description="Curso de prueba issue #58 A3",
            category=self.category,
            status=Course.Status.PUBLISHED,
            created_by=self.coordinador,
        )
        self.enrollment = Enrollment.objects.create(
            user=self.ejecutor,
            course=self.course,
            status=Enrollment.Status.IN_PROGRESS,
            progress=40,
        )

    def _headers_and_row(self, content):
        from io import BytesIO

        import openpyxl

        wb = openpyxl.load_workbook(BytesIO(content))
        ws = wb["Pendientes"]
        rows = list(ws.iter_rows(values_only=True))
        return rows[0], rows[1:]

    def test_export_incluye_columnas_rol_y_supervisor(self):
        content = ExportService.export_pending_users()
        headers, data_rows = self._headers_and_row(content)

        self.assertIn("Rol de Acceso", headers)
        self.assertIn("Documento Supervisor", headers)

        rol_idx = headers.index("Rol de Acceso")
        sup_idx = headers.index("Documento Supervisor")
        cedula_idx = headers.index("Cédula")

        matching = [r for r in data_rows if r[cedula_idx] == "600000201"]
        self.assertEqual(len(matching), 1)
        self.assertEqual(matching[0][rol_idx], "Ejecutor")
        self.assertEqual(matching[0][sup_idx], "600000200")

    def test_export_usuario_sin_rol_ni_supervisor_no_falla(self):
        """Usuario legacy sin `rol` asignado (bucket "sin asignar" del
        backfill de A1, o creado antes de A1) no debe romper el export."""
        sin_rol = User.objects.create_user(
            email="sin_rol_export_a3@example.com",
            password="x",
            first_name="SinRol",
            last_name="Export",
            document_type="CC",
            document_number="600000202",
            job_position="Liniero",
            job_profile=self.jp_liniero,
            hire_date=date(2023, 1, 1),
        )
        Enrollment.objects.create(
            user=sin_rol,
            course=self.course,
            status=Enrollment.Status.ENROLLED,
            progress=0,
        )

        content = ExportService.export_pending_users()
        headers, data_rows = self._headers_and_row(content)
        rol_idx = headers.index("Rol de Acceso")
        sup_idx = headers.index("Documento Supervisor")
        cedula_idx = headers.index("Cédula")

        matching = [r for r in data_rows if r[cedula_idx] == "600000202"]
        self.assertEqual(len(matching), 1)
        self.assertEqual(matching[0][rol_idx], "Sin asignar")
        # openpyxl relee una celda con cadena vacía como None; ambos
        # representan "sin supervisor asignado" para este export.
        self.assertFalse(matching[0][sup_idx])
