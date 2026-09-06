"""
Services for accounts app - password generation, bulk upload, and export.
"""

import io
import logging
from datetime import date

from django.contrib.auth import get_user_model
from django.db import IntegrityError, transaction

logger = logging.getLogger(__name__)
User = get_user_model()


class PasswordService:
    """Service for parameterized password generation."""

    @staticmethod
    def generate_password(document_number: str, first_name: str) -> str:
        """
        Generate a standardized password from user data.

        Format: document_number + first 3 letters of first_name (UPPERCASE)
        Example: document=1234567890, name=Carlos -> "1234567890CAR"
        """
        name_part = first_name.strip().upper()[:3] if first_name else "USR"
        # Pad if name is shorter than 3 chars
        name_part = name_part.ljust(3, "X")
        return f"{document_number}{name_part}"

    @staticmethod
    def reset_password(user) -> str:
        """Reset user password to the parameterized default."""
        new_password = PasswordService.generate_password(user.document_number, user.first_name)
        user.set_password(new_password)
        user.save(update_fields=["password"])
        logger.info(f"Password reset for user {user.document_number}")
        return new_password


class BulkUploadService:
    """Service for bulk user creation from Excel files."""

    REQUIRED_COLUMNS = [
        "nombre",
        "apellido",
        "numero_documento",
    ]

    # Aliases naturales que el usuario puede escribir a mano (o reutilizar de
    # un export previo) en vez del nombre exacto de columna. Se aplican DESPUÉS
    # de la normalización de acentos/espacios, antes de validar REQUIRED_COLUMNS.
    HEADER_ALIASES = {
        "cedula": "numero_documento",
        "documento": "numero_documento",
    }

    OPTIONAL_COLUMNS = [
        "tipo_documento",
        "correo",
        "telefono",
        "cargo",
        "perfil_ocupacional",
        "tipo_vinculacion",
        "fecha_ingreso",
        "estado",
        # RBAC (issue #58, sub-item A3) — ambas opcionales, ver ROL_MAP y
        # ROL_SUGERIDO_POR_JOB_PROFILE_CODE más abajo para el fallback.
        "rol_acceso",
        "supervisor_documento",
    ]

    DOCUMENT_TYPE_MAP = {
        "cc": "CC",
        "cédula de ciudadanía": "CC",
        "cedula de ciudadania": "CC",
        "ce": "CE",
        "cédula de extranjería": "CE",
        "ti": "TI",
        "tarjeta de identidad": "TI",
        "pasaporte": "PA",
        "pa": "PA",
    }

    JOB_PROFILE_MAP = {
        "liniero": "LINIERO",
        "técnico": "TECNICO",
        "tecnico": "TECNICO",
        "operador": "OPERADOR",
        "jefe de cuadrilla": "JEFE_CUADRILLA",
        "ingeniero residente": "INGENIERO_RESIDENTE",
        "ingeniero": "INGENIERO_RESIDENTE",
        "coordinador hseq": "COORDINADOR_HSEQ",
        "coordinador": "COORDINADOR_HSEQ",
        "administrador": "ADMINISTRADOR",
    }

    # Alias "amigables" arriba, para los 7 códigos con nombre común en
    # español. Los códigos SIN alias (CAPATAZ, "TODOS LOS CARGOS",
    # CONTRATISTA, CONDUCTOR, COORDINADOR_VIZ — issue #58, evidencia F2)
    # también deben poder crearse por bulk upload: la resolución de
    # `profile_code` en `create_users_from_rows()` hace fallback al código
    # literal (mayúsculas) tal como está en `job_profile_types.code` cuando
    # `perfil_ocupacional` no matchea ningún alias de este dict. Sin ese
    # fallback esos 5 códigos serían inalcanzables desde el Excel (se
    # degradaban en silencio a LINIERO).

    # Mapeo rol sugerido según `job_profile.code` — SOLO los 7 códigos con
    # sugerencia confiable (decisión de Miguel, HITL 2026-07-06). Es una
    # copia INTENCIONAL del mismo dict en `forms.py::RolSupervisorMixin`
    # (y conceptualmente el mismo mapeo de la migración de datos
    # `0016_backfill_user_rol.py`) — NO se importa desde ninguno de esos dos
    # lugares (importar un módulo de migración numerado desde código de app
    # es frágil; forms.py ya sentó el precedente de duplicar en vez de
    # importar). Si Miguel ajusta el mapeo, hay que tocar los 3 lugares.
    # A diferencia del backfill (que defaultea lo no-mapeado a EJECUTOR
    # para no dejar usuarios existentes sin rol), acá — igual que en
    # forms.py — lo no-mapeado NO tiene sugerencia: es alta hacia adelante,
    # se exige elección explícita vía la columna `rol_acceso`.
    ROL_SUGERIDO_POR_JOB_PROFILE_CODE = {
        "LINIERO": "EJECUTOR",
        "TECNICO": "EJECUTOR",
        "OPERADOR": "EJECUTOR",
        "JEFE_CUADRILLA": "COORDINADOR",
        "INGENIERO_RESIDENTE": "COORDINADOR",
        "COORDINADOR_HSEQ": "COORDINADOR",
        "ADMINISTRADOR": "ADMINISTRADOR",
    }

    # Valores aceptados en la columna opcional `rol_acceso` (case-insensitive,
    # coincide con los 3 `User.Rol` reales). Si la columna trae un valor que
    # NO está acá, es un error de fila (no se adivina un typo silenciosamente).
    ROL_MAP = {
        "ejecutor": "EJECUTOR",
        "coordinador": "COORDINADOR",
        "administrador": "ADMINISTRADOR",
    }

    EMPLOYMENT_TYPE_MAP = {
        "directo": "direct",
        "direct": "direct",
        "contratista": "contractor",
        "contractor": "contractor",
    }

    STATUS_MAP = {
        "activo": "active",
        "active": "active",
        "inactivo": "inactive",
        "inactive": "inactive",
        "suspendido": "suspended",
        "suspended": "suspended",
        "período de prueba": "probation",
        "periodo de prueba": "probation",
        "probation": "probation",
    }

    @staticmethod
    def parse_excel(file) -> tuple[list[dict], list[str]]:
        """
        Parse an Excel file and return rows and errors.
        Returns (rows, errors) where rows is a list of dicts.
        """
        try:
            import openpyxl
        except ImportError:
            return [], ["El paquete 'openpyxl' no está instalado."]

        errors = []

        try:
            wb = openpyxl.load_workbook(file, read_only=True, data_only=True)
            ws = wb.active
        except Exception as e:
            return [], [f"Error al leer el archivo Excel: {e}"]

        # Read headers from first row
        headers = []
        for cell in next(ws.iter_rows(min_row=1, max_row=1)):
            val = str(cell.value or "").strip().lower()
            # Normalize header names
            val = (
                val.replace("á", "a")
                .replace("é", "e")
                .replace("í", "i")
                .replace("ó", "o")
                .replace("ú", "u")
                .replace("ñ", "n")
                .replace(" ", "_")
            )
            val = BulkUploadService.HEADER_ALIASES.get(val, val)
            headers.append(val)

        # Validate required columns
        for col in BulkUploadService.REQUIRED_COLUMNS:
            normalized = (
                col.replace("á", "a")
                .replace("é", "e")
                .replace("í", "i")
                .replace("ó", "o")
                .replace("ú", "u")
            )
            if normalized not in headers:
                errors.append(f"Columna requerida '{col}' no encontrada en el archivo.")

        if errors:
            wb.close()
            return [], errors

        rows = []
        for row_num, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            if all(v is None for v in row):
                continue  # Skip empty rows

            row_data = {}
            for i, value in enumerate(row):
                if i < len(headers):
                    row_data[headers[i]] = value

            row_data["_row_num"] = row_num
            rows.append(row_data)

        wb.close()
        return rows, errors

    @staticmethod
    @transaction.atomic
    def create_users_from_rows(rows: list[dict]) -> tuple[list, list[str]]:
        """
        Create users from parsed rows.
        Returns (created_users, errors).
        """
        created = []
        errors = []

        for row in rows:
            row_num = row.get("_row_num", "?")

            # Required fields
            first_name = str(row.get("nombre", "")).strip()
            last_name = str(row.get("apellido", "")).strip()
            document_number = str(row.get("numero_documento", "")).strip()

            if not first_name or not last_name or not document_number:
                errors.append(f"Fila {row_num}: Nombre, apellido y documento son obligatorios.")
                continue

            # Check if user already exists
            if User.objects.filter(document_number=document_number).exists():
                errors.append(
                    f"Fila {row_num}: Ya existe un usuario con documento {document_number}."
                )
                continue

            # Optional fields
            doc_type_raw = str(row.get("tipo_documento", "CC")).strip().lower()
            document_type = BulkUploadService.DOCUMENT_TYPE_MAP.get(doc_type_raw, "CC")

            email = str(row.get("correo", "") or "").strip() or None
            phone = str(row.get("telefono", "") or "").strip()
            job_position = str(row.get("cargo", "Operario") or "Operario").strip()

            profile_raw_input = str(row.get("perfil_ocupacional", "LINIERO") or "LINIERO").strip()
            profile_raw = profile_raw_input.lower()
            if profile_raw in BulkUploadService.JOB_PROFILE_MAP:
                profile_code = BulkUploadService.JOB_PROFILE_MAP[profile_raw]
            else:
                # Sin alias amigable conocido: acepta el código real tal como
                # está en `job_profile_types.code` (ej. "CAPATAZ", "TODOS LOS
                # CARGOS", "COORDINADOR_VIZ", "CONTRATISTA", "CONDUCTOR" —
                # issue #58 A3, evidencia F2). Si tampoco matchea ningún
                # código real en BD, el fallback de abajo (`if not
                # job_profile: get(code="LINIERO")`) lo degrada a LINIERO,
                # igual que el comportamiento previo a A3.
                profile_code = profile_raw_input.upper() if profile_raw_input else "LINIERO"

            emp_raw = str(row.get("tipo_vinculacion", "directo") or "directo").strip().lower()
            employment_type = BulkUploadService.EMPLOYMENT_TYPE_MAP.get(emp_raw, "direct")

            status_raw = str(row.get("estado", "activo") or "activo").strip().lower()
            status = BulkUploadService.STATUS_MAP.get(status_raw, "active")

            # Parse hire_date
            hire_date_raw = row.get("fecha_ingreso")
            hire_date = hire_date_raw if isinstance(hire_date_raw, date) else date.today()

            # Check email uniqueness
            if email and User.objects.filter(email=email).exists():
                errors.append(f"Fila {row_num}: Ya existe un usuario con correo {email}.")
                continue

            from apps.courses.models import JobProfileType

            job_profile = JobProfileType.objects.filter(code=profile_code).first()
            if not job_profile:
                job_profile = JobProfileType.objects.get(code="LINIERO")

            # --- Rol de acceso (RBAC, issue #58 A3) ---
            # Misma regla de A2 (`RolSupervisorMixin`): valor explícito de la
            # columna `rol_acceso` gana; si viene vacío, se aplica la
            # sugerencia según `job_profile.code` SOLO si es confiable (7
            # códigos mapeados); si no hay sugerencia confiable, se exige
            # elección explícita — la fila se rechaza, no se adivina un rol.
            rol_raw = str(row.get("rol_acceso", "") or "").strip().lower()
            if rol_raw:
                rol = BulkUploadService.ROL_MAP.get(rol_raw)
                if rol is None:
                    errors.append(
                        f"Fila {row_num}: valor de 'rol_acceso' no reconocido "
                        f"('{row.get('rol_acceso')}'). Use Ejecutor, "
                        "Coordinador o Administrador."
                    )
                    continue
            else:
                rol = BulkUploadService.ROL_SUGERIDO_POR_JOB_PROFILE_CODE.get(job_profile.code)
                if rol is None:
                    errors.append(
                        f"Fila {row_num}: el perfil ocupacional "
                        f"'{job_profile.code}' no sugiere un rol de acceso "
                        "automáticamente; indique la columna 'rol_acceso' "
                        "explícitamente (Ejecutor, Coordinador o "
                        "Administrador)."
                    )
                    continue

            # --- Supervisor (FK auto-referencial, issue #58 A3) ---
            # Misma regla de A2 (`User.clean()` / `RolSupervisorMixin`):
            # rechaza auto-supervisión y exige que el supervisor exista y
            # tenga rol Coordinador o Administrador (mismo queryset que
            # restringe el selector del form en A2).
            supervisor_doc = str(row.get("supervisor_documento", "") or "").strip()
            supervisor = None
            if supervisor_doc:
                if supervisor_doc == document_number:
                    errors.append(
                        f"Fila {row_num}: un usuario no puede ser supervisor "
                        "de sí mismo ('supervisor_documento' coincide con "
                        "'numero_documento')."
                    )
                    continue
                supervisor = User.objects.filter(document_number=supervisor_doc).first()
                if supervisor is None:
                    errors.append(
                        f"Fila {row_num}: no existe un usuario con documento "
                        f"'{supervisor_doc}' para asignar como supervisor."
                    )
                    continue
                if supervisor.rol not in (User.Rol.COORDINADOR, User.Rol.ADMINISTRADOR):
                    supervisor_rol_display = (
                        supervisor.get_rol_display() if supervisor.rol else "Sin asignar"
                    )
                    errors.append(
                        f"Fila {row_num}: el supervisor con documento "
                        f"'{supervisor_doc}' debe tener rol Coordinador o "
                        f"Administrador (tiene '{supervisor_rol_display}')."
                    )
                    continue

            # Generate password
            password = PasswordService.generate_password(document_number, first_name)

            try:
                with transaction.atomic():
                    user = User(
                        email=email,
                        first_name=first_name,
                        last_name=last_name,
                        document_type=document_type,
                        document_number=document_number,
                        phone=phone,
                        job_position=job_position,
                        job_profile=job_profile,
                        rol=rol,
                        supervisor=supervisor,
                        employment_type=employment_type,
                        hire_date=hire_date,
                        status=status,
                        is_active=status == "active",
                    )
                    user.set_password(password)
                    user.save()
                # Atributo transiente (NO campo de modelo): expone la
                # contraseña generada para que la plantilla la muestre en la
                # tabla de resultados — sin esto, la carga masiva creaba
                # cuentas reales cuyo password nadie veía nunca (issue #58,
                # reproceso).
                user.generated_password = password
                created.append(user)
            except IntegrityError:
                logger.exception(f"Fila {row_num}: IntegrityError al crear usuario")
                errors.append(
                    f"Fila {row_num}: No se pudo crear el usuario debido a un error "
                    "interno. Contacte al administrador."
                )
            except Exception:
                logger.exception(f"Fila {row_num}: Error inesperado al crear usuario")
                errors.append(
                    f"Fila {row_num}: No se pudo crear el usuario debido a un error "
                    "interno. Contacte al administrador."
                )

        logger.info(f"Bulk upload: {len(created)} created, {len(errors)} errors")
        return created, errors

    @staticmethod
    def generate_template() -> bytes:
        """Generate a template Excel file for bulk upload."""
        try:
            import openpyxl
        except ImportError:
            return b""

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Usuarios"

        headers = [
            "nombre",
            "apellido",
            "numero_documento",
            "tipo_documento",
            "correo",
            "telefono",
            "cargo",
            "perfil_ocupacional",
            "tipo_vinculacion",
            "fecha_ingreso",
            "estado",
            "rol_acceso",
            "supervisor_documento",
        ]
        ws.append(headers)

        # Example row
        ws.append(
            [
                "Juan",
                "Pérez",
                "1234567890",
                "CC",
                "juan@ejemplo.com",
                "3001234567",
                "Técnico Electricista",
                "TECNICO",
                "directo",
                date.today().isoformat(),
                "activo",
                "",  # rol_acceso vacío -> se sugiere EJECUTOR por TECNICO
                "",  # supervisor_documento vacío -> sin supervisor asignado
            ]
        )

        # Set column widths
        for col_idx, header in enumerate(headers, 1):
            ws.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = 20

        # Second sheet documenting conventions (column aliases + password rule)
        notes_ws = wb.create_sheet("Instrucciones")
        notes_ws.append(["Columna / Regla", "Descripción"])
        notes_ws.append(
            [
                "numero_documento",
                "También se acepta 'Cédula' o 'documento' como nombre de columna.",
            ]
        )
        notes_ws.append(
            [
                "Contraseña generada",
                "numero_documento + 3 primeras letras del nombre en MAYÚSCULA "
                "(ej: documento 1234567890 + nombre Carlos -> 1234567890CAR).",
            ]
        )
        notes_ws.append(
            [
                "rol_acceso (opcional)",
                "Rol de acceso RBAC: Ejecutor, Coordinador o Administrador "
                "(no distingue mayúsculas/minúsculas). Si se deja vacío, se "
                "sugiere automáticamente según 'perfil_ocupacional' cuando "
                "este tiene un mapeo confiable (Liniero/Técnico/Operador -> "
                "Ejecutor; Jefe de Cuadrilla/Ingeniero Residente/Coordinador "
                "HSEQ -> Coordinador; Administrador -> Administrador). Si el "
                "perfil no sugiere un rol confiable (ej. Capataz, Contratista, "
                "Conductor, Coordinador Solo Visualización, Todos los Cargos), "
                "la fila se rechaza si esta columna queda vacía.",
            ]
        )
        notes_ws.append(
            [
                "supervisor_documento (opcional)",
                "Número de documento de un usuario YA EXISTENTE con rol "
                "Coordinador o Administrador. Si se deja vacío, el usuario "
                "queda sin supervisor asignado. La fila se rechaza si el "
                "documento no existe, si el usuario referido no tiene rol "
                "Coordinador/Administrador, o si coincide con el propio "
                "'numero_documento' de la fila (auto-supervisión).",
            ]
        )
        notes_ws.column_dimensions["A"].width = 25
        notes_ws.column_dimensions["B"].width = 70

        output = io.BytesIO()
        wb.save(output)
        return output.getvalue()


class ExportService:
    """Service for exporting data to Excel."""

    @staticmethod
    def export_pending_users(category_id=None, profile=None) -> bytes:
        """
        Export users with pending/incomplete courses to Excel.
        """
        try:
            import openpyxl
        except ImportError:
            return b""

        from apps.courses.models import Enrollment

        # Get active enrollments that are NOT completed
        enrollments = (
            Enrollment.objects.filter(
                status__in=[
                    Enrollment.Status.ENROLLED,
                    Enrollment.Status.IN_PROGRESS,
                    Enrollment.Status.EXPIRED,
                ]
            )
            .select_related("user", "user__supervisor", "course", "course__category")
            .order_by("user__last_name", "user__first_name", "course__title")
        )

        # Apply filters
        if category_id:
            enrollments = enrollments.filter(course__category_id=category_id)

        if profile:
            enrollments = enrollments.filter(user__job_profile__code=profile)

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Pendientes"

        headers = [
            "Nombre",
            "Apellido",
            "Cédula",
            "Perfil Ocupacional",
            "Cargo",
            "Rol de Acceso",
            "Documento Supervisor",
            "Curso",
            "Categoría",
            "Estado",
            "Progreso (%)",
            "Fecha Inscripción",
            "Fecha Límite",
        ]
        ws.append(headers)

        # Style headers
        for col_idx in range(1, len(headers) + 1):
            cell = ws.cell(row=1, column=col_idx)
            cell.font = openpyxl.styles.Font(bold=True)
            ws.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = 20

        status_display = {
            "enrolled": "Inscrito",
            "in_progress": "En progreso",
            "expired": "Vencido",
        }

        for enrollment in enrollments:
            ws.append(
                [
                    enrollment.user.first_name,
                    enrollment.user.last_name,
                    enrollment.user.document_number,
                    enrollment.user.job_profile.name if enrollment.user.job_profile else "N/A",
                    enrollment.user.job_position,
                    enrollment.user.get_rol_display() if enrollment.user.rol else "Sin asignar",
                    enrollment.user.supervisor.document_number
                    if enrollment.user.supervisor_id
                    else "",
                    enrollment.course.title,
                    enrollment.course.category.name
                    if enrollment.course.category
                    else "Sin categoría",
                    status_display.get(enrollment.status, enrollment.status),
                    float(enrollment.progress),
                    enrollment.created_at.strftime("%Y-%m-%d") if enrollment.created_at else "",
                    enrollment.due_date.strftime("%Y-%m-%d")
                    if enrollment.due_date
                    else "Sin fecha",
                ]
            )

        output = io.BytesIO()
        wb.save(output)
        return output.getvalue()
